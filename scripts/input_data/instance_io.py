"""
Input data utilities for aircraft positioning instances.

Functions
---------
validate_instance(data)         Validate a dict against instance_schema.json.
load_json(filepath)             Read a JSON instance file, validate it, return dict.
read_xlsx(xlsx_path)            Read an xlsx file and return the raw instance dict.
xlsx_to_json(xlsx_path, ...)    Convert an xlsx file to a JSON file on disk.
convert_all_scenarios(...)      Batch-convert every xlsx in a scenarios directory.

xlsx columns expected
---------------------
  plane    : integer plane ID        → "R{n}"
  task     : task order within plane (1 = first, max = last)
  job      : job ID string           → "J{id}"
  duration : task duration
  client   : integer client ID       → "C{n}"
  es       : earliest start (first task of plane)
  lf       : latest finish  (last  task of plane)

Hangar layout (fixed)
---------------------
  Positions    : P1, P2, P3, P4, P5
  Blocking arcs: P3→P4, P3→P5, P4→P5
"""

import json
import sys
from pathlib import Path

import jsonschema
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

HANGAR = {
    "positions": ["P1", "P2", "P3", "P4", "P5"],
    "blocking_arcs": [
        {"front": "P3", "rear": "P4"},
        {"front": "P3", "rear": "P5"},
        {"front": "P4", "rear": "P5"},
    ],
}

_SCHEMA_PATH  = Path(__file__).resolve().parent / "instance_schema.json"
OUTPUT_DIR    = Path(__file__).resolve().parents[2] / "data"
SCENARIOS_DIR = Path("c:/Users/alvaro/repos/aircraft-positioning-data/input/scenarios")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _aircraft_id(plane: int) -> str:
    return f"R{plane}"

def _client_id(client: int) -> str:
    return f"C{client}"

def _job_id(raw: str) -> str:
    return f"J{raw}"

# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate_instance(data: dict) -> None:
    """Validate *data* against instance_schema.json.

    Raises
    ------
    ValueError
        If *data* does not conform to the schema.
    """
    with open(_SCHEMA_PATH) as f:
        schema = json.load(f)
    try:
        jsonschema.validate(instance=data, schema=schema)
    except jsonschema.ValidationError as exc:
        raise ValueError(f"Instance does not conform to schema: {exc.message}") from exc

# ---------------------------------------------------------------------------
# JSON loading
# ---------------------------------------------------------------------------

def load_json(filepath: str | Path) -> dict:
    """Read a JSON instance file, validate it, and return the data dict.

    Raises
    ------
    ValueError
        If the file does not conform to the instance schema.
    """
    with open(filepath, encoding="utf-8") as f:
        data = json.load(f)
    validate_instance(data)
    return data

# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------

def read_xlsx(xlsx_path: str | Path) -> dict:
    """Read an xlsx scenario file and return a validated instance dict.

    The returned dict has the same structure produced by load_json.

    Raises
    ------
    ValueError
        If the resulting instance does not conform to the schema.
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    planes: dict[int, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        plane = int(row[col["plane"]])
        planes.setdefault(plane, []).append({
            "task":     int(row[col["task"]]),
            "job":      str(row[col["job"]]),
            "duration": float(row[col["duration"]]),
            "client":   int(row[col["client"]]),
            "es":       float(row[col["es"]]),
            "lf":       float(row[col["lf"]]),
        })

    for tasks in planes.values():
        tasks.sort(key=lambda r: r["task"])

    aircrafts = []
    for plane, tasks in sorted(planes.items()):
        aircrafts.append({
            "id":             _aircraft_id(plane),
            "client":         _client_id(tasks[0]["client"]),
            "earliest_start": tasks[0]["es"],
            "target_finish":  tasks[-1]["lf"],
        })

    jobs = []
    for plane, tasks in sorted(planes.items()):
        max_task = tasks[-1]["task"]
        for t in tasks:
            jobs.append({
                "id":          _job_id(t["job"]),
                "aircraft_id": _aircraft_id(plane),
                "duration":    t["duration"],
                "is_first":    t["task"] == 1,
                "is_last":     t["task"] == max_task,
            })

    job_precedences = []
    for plane, tasks in sorted(planes.items()):
        for i in range(len(tasks) - 1):
            job_precedences.append({
                "before": _job_id(tasks[i]["job"]),
                "after":  _job_id(tasks[i + 1]["job"]),
            })

    data = {
        # Default tow-in/tow-out time (epsilon); xlsx scenarios do not carry
        # this value, so we fall back to the convention used in the paper.
        "min_separation": 0.5,
        "hangar":          HANGAR,
        "aircrafts":       aircrafts,
        "jobs":            jobs,
        "job_precedences": job_precedences,
    }
    validate_instance(data)
    return data

# ---------------------------------------------------------------------------
# xlsx → JSON file
# ---------------------------------------------------------------------------

def xlsx_to_json(
    xlsx_path: str | Path,
    output_dir: str | Path = OUTPUT_DIR,
) -> Path:
    """Convert an xlsx scenario file to a JSON instance file on disk.

    Parameters
    ----------
    xlsx_path:
        Path to the source xlsx file.
    output_dir:
        Directory where the JSON file will be written.
        Defaults to the project ``data/`` folder.

    Returns
    -------
    Path
        Path to the written JSON file.
    """
    xlsx_path  = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    data = read_xlsx(xlsx_path)
    output_path = output_dir / (xlsx_path.stem + ".json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"Written: {output_path}")
    return output_path

# ---------------------------------------------------------------------------
# Batch conversion
# ---------------------------------------------------------------------------

def convert_all_scenarios(
    scenarios_dir: str | Path = SCENARIOS_DIR,
    output_dir: str | Path = OUTPUT_DIR,
) -> list[Path]:
    """Convert every xlsx file in *scenarios_dir* to a JSON instance file.

    Parameters
    ----------
    scenarios_dir:
        Directory containing xlsx scenario files.
    output_dir:
        Directory where JSON files will be written.

    Returns
    -------
    list[Path]
        Paths of the JSON files that were written.
    """
    scenarios_dir = Path(scenarios_dir)
    xlsx_files = sorted(scenarios_dir.glob("*.xlsx"))

    if not xlsx_files:
        print(f"No xlsx files found in: {scenarios_dir}", file=sys.stderr)
        return []

    written = []
    for xlsx_path in xlsx_files:
        try:
            written.append(xlsx_to_json(xlsx_path, output_dir))
        except Exception as exc:
            print(f"Error processing {xlsx_path.name}: {exc}", file=sys.stderr)

    print(f"\n{len(written)}/{len(xlsx_files)} files converted successfully.")
    return written

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) == 1:
        # No arguments → batch convert all scenarios
        convert_all_scenarios()
    elif len(sys.argv) == 2:
        xlsx_path = Path(sys.argv[1])
        if not xlsx_path.exists():
            print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
            sys.exit(1)
        xlsx_to_json(xlsx_path)
    else:
        print("Usage: instance_io.py [<xlsx_file>]", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
