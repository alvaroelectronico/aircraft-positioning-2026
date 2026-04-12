"""
Reads an xlsx file and converts it to a JSON instance file conforming to instance_schema.json.

xlsx columns:
  plane    : integer plane ID  → converted to "R{n}"
  task     : task order within the plane (1 = first, max = last)
  job      : job ID string from the xlsx  → prefixed as "J{id}"
  duration : task duration
  client   : integer client ID  → converted to "C{n}"
  es       : earliest start (only meaningful for the first task of a plane)
  lf       : latest finish   (only meaningful for the last  task of a plane)

Hangar layout (fixed):
  Positions   : P1, P2, P3, P4, P5
  Blocking arcs: (P3→P4), (P3→P5), (P4→P5)

Output: data/<xlsx_stem>.json
"""

import json
import sys
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------


HANGAR = {
    "positions": ["P1", "P2", "P3", "P4", "P5"],
    "blocking_arcs": [
        {"front": "P3", "rear": "P4"},
        {"front": "P3", "rear": "P5"},
        {"front": "P4", "rear": "P5"},
    ],
}

OUTPUT_DIR = Path(__file__).resolve().parents[2] / "data"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def aircraft_id(plane: int) -> str:
    return f"R{plane}"

def client_id(client: int) -> str:
    return f"C{client}"

def job_id(raw: str) -> str:
    return f"J{raw}"

# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def xlsx_to_instance(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    # Group rows by plane, preserving task order
    planes: dict[int, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        plane  = int(row[col["plane"]])
        task   = int(row[col["task"]])
        planes.setdefault(plane, []).append({
            "task":     task,
            "job":      str(row[col["job"]]),
            "duration": float(row[col["duration"]]),
            "client":   int(row[col["client"]]),
            "es":       float(row[col["es"]]),
            "lf":       float(row[col["lf"]]),
        })

    # Sort tasks within each plane by task number
    for tasks in planes.values():
        tasks.sort(key=lambda r: r["task"])

    # Build aircrafts
    aircrafts = []
    for plane, tasks in sorted(planes.items()):
        first_task = tasks[0]
        last_task  = tasks[-1]
        aircrafts.append({
            "id":             aircraft_id(plane),
            "client":         client_id(first_task["client"]),
            "earliest_start": first_task["es"],
            "target_finish":  last_task["lf"],
        })

    # Build jobs
    jobs = []
    for plane, tasks in sorted(planes.items()):
        max_task = tasks[-1]["task"]
        for task_row in tasks:
            jobs.append({
                "id":          job_id(task_row["job"]),
                "aircraft_id": aircraft_id(plane),
                "duration":    task_row["duration"],
                "is_first":    task_row["task"] == 1,
                "is_last":     task_row["task"] == max_task,
            })

    # Build job precedences (consecutive tasks within each plane)
    job_precedences = []
    for plane, tasks in sorted(planes.items()):
        for i in range(len(tasks) - 1):
            job_precedences.append({
                "before": job_id(tasks[i]["job"]),
                "after":  job_id(tasks[i + 1]["job"]),
            })

    return {
        "hangar":           HANGAR,
        "aircrafts":        aircrafts,
        "jobs":             jobs,
        "job_precedences":  job_precedences,
    }


def main():
    # ---- cambia esta ruta para probar distintas instancias ----
    default_xlsx = Path(
        "c:/Users/alvaro/repos/aircraft-positioning-data/input/scenarios/"
        "scn_few-loose_seed1_P3_pl5.xlsx"
    )
    # -----------------------------------------------------------
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx

    if not xlsx_path.exists():
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    instance = xlsx_to_instance(xlsx_path)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / (xlsx_path.stem + ".json")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(instance, f, indent=2, ensure_ascii=False)

    print(f"Instance written to: {output_path}")


if __name__ == "__main__":
    main()
